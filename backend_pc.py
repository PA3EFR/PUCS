#!/usr/bin/env python3
"""
Pile-Up COntrol System (PUCS) - Flask Backend (voor PC)
CORS-enabled voor frontend communicatie van pucs.pa3efr.nl
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date
import os
from flask_socketio import SocketIO, emit
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import threading
import time
import requests
import html
import urllib.parse
import sqlite3
from contextlib import contextmanager
import re
import atexit
import signal
import sys

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")  # CORS voor frontend

# CORS configuratie voor frontend communicatie
CORS(app, resources={
    r"/api/*": {
        "origins": [
            "https://pucs.pa3efr.nl",
            "http://pucs.pa3efr.nl",
            "http://localhost:*",
            "http://127.0.0.1:*"
        ],
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Accept"]
    },
    r"/admin/*": {
        "origins": [
            "https://pucs.pa3efr.nl",
            "http://pucs.pa3efr.nl",
            "http://localhost:*",
            "http://127.0.0.1:*"
        ],
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Accept"]
    }
})

app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///radio_entry.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

class Config(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    operator_name = db.Column(db.String(100), nullable=False, default='Operator')
    frequency = db.Column(db.String(50), nullable=False, default='145.500 MHz')
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

    @staticmethod
    def get_current():
        config = Config.query.first()
        if not config:
            config = Config(operator_name='QRT for now', frequency='---')
            db.session.add(config)
            db.session.commit()
        return config

class CallsignEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    position = db.Column(db.Integer, nullable=False, unique=True)  # 1-4
    callsign = db.Column(db.String(20), nullable=False)
    location = db.Column(db.String(100), nullable=True)  # QTH field toegevoegd
    comment = db.Column(db.Text, nullable=True)  # Remarks field toegevoegd
    entered_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'position': self.position,
            'callsign': self.callsign,
            'location': self.location,
            'comment': self.comment,
            'entered_at': self.entered_at.isoformat()
        }

class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class QRZConfig(db.Model):
    """QRZ API configuratie tabel"""
    __tablename__ = 'qrz_config'
    
    id = db.Column(db.Integer, primary_key=True)
    callsign = db.Column(db.String(20), nullable=False)
    api_key = db.Column(db.String(100), nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    @staticmethod
    def get_current():
        """Haal huidige QRZ configuratie op"""
        config = QRZConfig.query.order_by(QRZConfig.updated_at.desc()).first()
        return config

class QRZLogbookChecker:
    """
    QRZ Logbook Checker - Geïntegreerd in backend
    Controleert elke minuut het QRZ logboek en verwijdert gelogde callsigns van vandaag uit de entry list
    """
    
    def __init__(self, db_path='instance/radio_entry.db'):
        self.db_path = db_path
        self.running = False
        self.check_thread = None
        self.interval = 60  # 1 minuten in seconden

    def log(self, message):
        """Minimal logging: alleen belangrijke meldingen tonen"""
        allowed_prefixes = [
            "🚀 QRZ Logbook Checker gestart",
            "✅ Fetch",
            "🎯 Laatste callsign gevonden"
        ]
        if any(message.startswith(prefix) for prefix in allowed_prefixes):
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[{timestamp}] QRZ Checker: {message}")
        
    @contextmanager
    def get_db_connection(self):
        """Database connectie context manager"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
        finally:
            conn.close()
            
    def get_qrz_config(self):
        """Haal QRZ configuratie op uit database"""
        try:
            with self.get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT callsign, api_key FROM qrz_config ORDER BY updated_at DESC LIMIT 1")
                row = cursor.fetchone()
                
                if row:
                    return row['callsign'], row['api_key']
                else:
                    # Fallback naar default waarden
                    self.log("Geen QRZ configuratie gevonden, gebruik default waarden")
                    return 'PH25XMAS', '4DB9-....-579F'
                    
        except Exception as e:
            self.log(f"Fout bij ophalen QRZ config: {e}")
            return 'PH25XMAS', '4DB9-...-579F'
            
    def get_today_active_callsigns(self):
        """Haal actieve callsigns van vandaag op uit database"""
        try:
            today = date.today()
            with self.get_db_connection() as conn:
                cursor = conn.cursor()
                # Alleen callsigns van vandaag ophalen
                cursor.execute("""
                    SELECT callsign FROM callsign_entry 
                    WHERE DATE(entered_at) = ?
                """, (today,))
                rows = cursor.fetchall()
                callsigns = [row['callsign'] for row in rows]
                
                if callsigns:
                    self.log(f"📋 {len(callsigns)} callsigns van vandaag ({today}) gevonden: {', '.join(callsigns)}")
                
                return callsigns
                
        except Exception as e:
            self.log(f"Fout bij ophalen actieve callsigns van vandaag: {e}")
            return []
            
    def remove_callsign_from_db(self, callsign):
        """Verwijder ALLEEN de laatste entry van callsign uit database (alleen van vandaag)"""
        try:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            today = date.today()
            self.log(f"🐛 [{current_time}] Starting removal process for {callsign}...")
            
            with self.get_db_connection() as conn:
                cursor = conn.cursor()
                
                # Zoek de laatste entry van deze callsign van vandaag
                cursor.execute("""
                    SELECT id, entered_at FROM callsign_entry 
                    WHERE callsign = ? AND DATE(entered_at) = ?
                    ORDER BY entered_at DESC LIMIT 1
                """, (callsign, today))
                
                row = cursor.fetchone()
                if not row:
                    self.log(f"ℹ️ Geen entry van {callsign} van vandaag gevonden in database")
                    return False
                
                entry_id = row['id']
                entry_time = row['entered_at']
                
                self.log(f"🐛 Found latest entry: id={entry_id}, time={entry_time}")
                
                # Verwijder alleen deze specifieke entry
                cursor.execute("""
                    DELETE FROM callsign_entry 
                    WHERE id = ?
                """, (entry_id,))
                
                if cursor.rowcount > 0:
                    conn.commit()
                    self.log(f"✅ Laatste entry van {callsign} (id: {entry_id}, tijd: {entry_time}) verwijderd uit entry list")
                    return True
                else:
                    self.log(f"⚠️ Kon entry {entry_id} van {callsign} niet verwijderen")
                    return False
                
        except Exception as e:
            self.log(f"Fout bij verwijderen laatste entry van {callsign}: {e}")
            return False
            
    def fetch_qrz_logbook(self, callsign, api_key):
        """Haal QRZ logboek op via API met complete download strategie"""
        try:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.log(f"🔍 [{current_time}] COMPLETE QRZ LOGBOEK DOWNLOAD voor {callsign}...")
            self.log(f"🔑 Using QRZ API Key: {api_key[:8]}...{api_key[-8:]} (masked for security)")
            
            # Ultra-agressieve anti-caching headers
            headers = {
                'Cache-Control': 'no-cache, no-store, must-revalidate, private',
                'Pragma': 'no-cache',
                'Expires': '0',
                'If-Modified-Since': 'Wed, 21 Oct 2020 07:28:00 GMT',  # Force old date
                'If-None-Match': '"no-cache"',
                'User-Agent': 'Mozilla/5.0 (compatible; PUCS-QRZ-Checker/1.0)',
                'Accept': 'application/x-www-form-urlencoded'
            }
            
            # COMPLETE ADIF DOWNLOAD STRATEGIE
            # Probeer verschillende fetch configuraties om alle data te krijgen
            today_str = date.today().strftime('%Y-%m-%d')
            
            fetch_configs = [
                # Config 1: Alle records (base)
                {
                    "OPTION": "ALL",
                    "MAX": 1000,
                    "ts": str(int(datetime.now().timestamp() * 1000))
                },
                # Config 2: Datum range (october 2025)
                {
                    "OPTION": f"BETWEEN:2025-10-01+{today_str}",
                    "MAX": 1000,
                    "ts": str(int(datetime.now().timestamp() * 1000))
                },
                # Config 3: Modified since
                {
                    "OPTION": "MODSINCE:2025-10-01",
                    "MAX": 1000,
                    "ts": str(int(datetime.now().timestamp() * 1000))
                },
                # Config 4: Recent records only
                {
                    "OPTION": "ALL",
                    "MAX": 500,
                    "ts": str(int(datetime.now().timestamp() * 1000))
                }
            ]
            
            all_adif_data = ""  # Combineer alle responses
            successful_fetches = 0
            total_records_found = 0
            
            # Probeer alle fetch configuraties
            for i, config in enumerate(fetch_configs):
                try:
                    self.log(f"🔄 Fetch attempt {i+1}/4: OPTION={config.get('OPTION', 'ALL')}")
                    
                    params = {
                        "KEY": api_key,
                        "ACTION": "FETCH",
                        "ADIF": 1,
                        **config  # Merge config
                    }
                    
                    self.log(f"📡 API params: {params}")
                    
                    response = requests.get("https://logbook.qrz.com/api", params=params, headers=headers, timeout=30)
                    
                    if response.status_code == 200:
                        response_data = response.text
                        self.log(f"✅ Fetch {i+1} success: {len(response_data)} characters")
                        
                        # CRITICAL FIX: Decode HTML entities first
                        decoded_data = self.decode_html_entities(response_data)
                        self.log(f"🔧 Decoded response {i+1}: {len(decoded_data)} characters")
                        
                        # Parse ADIF om aantal records te tellen (gebruik decoded data)
                        records = self._parse_adif_records_count(decoded_data)
                        self.log(f"📊 Records in response {i+1}: {records}")
                        total_records_found += records
                        
                        # Combineer data (verwijder duplicaten) - gebruik decoded data
                        if decoded_data not in all_adif_data:
                            all_adif_data += decoded_data + "\n\n"
                            successful_fetches += 1
                        else:
                            self.log(f"ℹ️ Response {i+1} duplicate, skipping")
                    else:
                        self.log(f"❌ Fetch {i+1} failed: HTTP {response.status_code}")
                        
                except Exception as e:
                    self.log(f"⚠️ Fetch {i+1} error: {e}")
                    continue
            
            # Gebruik de best/alle data
            if all_adif_data:
                adif_data = all_adif_data.strip()
                self.log(f"🎯 COMPLETE ADIF DATA: {successful_fetches} successful fetches")
                self.log(f"📊 Total unique records detected: {total_records_found}")
                self.log(f"📋 Final ADIF size: {len(adif_data)} characters")
                
                # CRITICAL: Decode HTML entities first, then save
                decoded_adif = self.decode_html_entities(adif_data)
                
                # CRITICAL: Save complete ADIF immediately after fetch with metadata
                self.debug_save_adif_data_enhanced(
                    decoded_adif, 
                    context='complete_fetch',
                    metadata={
                        'successful_fetches': successful_fetches,
                        'total_records': total_records_found,
                        'callsign': callsign,
                        'api_key_masked': f"{api_key[:8]}...{api_key[-4:]}"
                    }
                )
                
                # === FIX: HTML ENTITY DECODING ===
                import urllib.parse
                from html import unescape
                
                try:
                    # Parse QRZ response als URL parameters
                    # QRZ format: RESULT=OK&COUNT=4&ADIF=<html_encoded_data>
                    response_dict = {}
                    adif_parts = []
                    
                    parts = adif_data.split('&')
                    for i, pair in enumerate(parts):
                        if '=' in pair:
                            key, value = pair.split('=', 1)
                            response_dict[key] = value
                        else:
                            # Dit is ADIF data na de eerste twee parameters
                            if i >= 2:  # Skip RESULT en COUNT
                                adif_parts.append(pair)
                    
                    # ADIF data is alles na RESULT=OK&COUNT=X
                    adif_encoded = '&'.join(adif_parts)
                    
                    # Extract en decode ADIF data
                    if adif_encoded:
                        
                        if adif_encoded:
                            self.log(f"📦 ADIF Raw Data: {len(adif_encoded)} characters")
                            self.log(f"📝 ADIF Preview: {adif_encoded[:100]}...")
                            
                            # Complete decodering keten
                            adif_decoded = urllib.parse.unquote_plus(adif_encoded)  # URL decode
                            adif_final = unescape(adif_decoded)  # HTML entities decode ← CRITICAL FIX!
                            
                            self.log(f"🔄 QRZ HTML entities decoded: {len(adif_encoded)} → {len(adif_final)} chars")
                            adif_data = adif_final  # Gebruik decoded data
                            
                            # Count callsigns na decodering
                            import re
                            calls = re.findall(r'<CALL:(\d+)>([A-Z0-9/]+)', adif_data, re.IGNORECASE)
                            self.log(f"📞 Callsigns na decodering: {len(calls)}")
                            if calls:
                                all_calls = [c[1] for c in calls]
                                self.log(f"📋 Callsigns gevonden: {all_calls}")
                        else:
                            self.log("⚠️ ADIF field is empty in QRZ response")
                    else:
                        self.log("⚠️ No ADIF field found in QRZ response - possible API response format issue")
                        
                except Exception as e:
                    self.log(f"⚠️ Error during QRZ response decoding: {e}")
                    # Gebruik originale data als fallback
                
                # === EINDE FIX ===
                
            else:
                self.log("❌ No data retrieved from any fetch attempt")
                return None
            
            self.log(f"📡 COMPLETE ADIF response succesvol ontvangen ({len(adif_data)} karakters)")
            
            # Debug preview
            adif_preview = adif_data[:300] if len(adif_data) > 300 else adif_data
            self.log(f"🐛 ADIF preview: {adif_preview}...")
            
            # Check ADIF data voor test/demo indicatoren
            if 'TE5T' in adif_data and 'test call' in adif_data.lower():
                self.log(f"⚠️ TEST DATA GEVONDEN: ADIF bevat test callsigns (TE5T)")
                self.log(f"ℹ️ INFO: Dit kan betekenen demo/limitatie data van QRZ.com")
                self.log(f"ℹ️ INFO: Controleer of je correct bent ingelogd bij QRZ.com")
            elif 'TE5T' in adif_data:
                self.log(f"ℹ️ INFO: ADIF bevat TE5T test callsign - controleer data inhoud")
            
            return adif_data
            
            # Check API credentials status (voor deze laatste fetch)
            if api_key == '4DB9-.....-579F' or not api_key:
                self.log(f"ℹ️ INFO: Using default/demo API key - this may limit data access")
                self.log(f"ℹ️ INFO: For full access, ensure proper QRZ.com login and API key configuration")
            else:
                self.log(f"✅ Using custom QRZ API key: {api_key[:8]}...{api_key[-4:]}")
            
            # Response processing (met already successful data)
            self.log(f"📡 COMPLETE ADIF response processing: {len(adif_data)} characters")
            
            # Check ADIF data voor demo/limitation indicators
            if '<call:4>TE5T' in adif_data and 'test call' in adif_data.lower():
                self.log(f"⚠️ TEST DATA GEVONDEN: ADIF bevat test callsigns (TE5T)")
                self.log(f"ℹ️ INFO: Dit kan betekenen demo/limitatie data van QRZ.com")
                self.log(f"ℹ️ INFO: Controleer of je correct bent ingelogd bij QRZ.com")
            elif 'TE5T' in adif_data:
                self.log(f"ℹ️ INFO: ADIF bevat TE5T test callsign - controleer data inhoud")
            
            self.log(f"📡 COMPLETE ADIF response succesvol ontvangen ({len(adif_data)} karakters)")
            
            # Debug preview
            adif_preview = adif_data[:300] if len(adif_data) > 300 else adif_data
            self.log(f"🐛 ADIF preview: {adif_preview}...")
            
            return adif_data
            
        except requests.exceptions.Timeout:
            self.log("❌ QRZ API timeout")
            return None
        except Exception as e:
            self.log(f"❌ Complete QRZ fetch error: {e}")
            return None
    
    def _parse_adif_records_count(self, adif_data):
        """Tel aantal QSO records in ADIF data"""
        try:
            # Count <eor> markers (end of record)
            eor_count = adif_data.count('<eor>')
            self.log(f"🔢 EOR markers found: {eor_count}")
            
            # Count <call: tags
            call_count = adif_data.count('<call:')
            self.log(f"🔢 CALL tags found: {call_count}")
            
            # Count <qso_date: tags  
            date_count = adif_data.count('<qso_date:')
            self.log(f"🔢 QSO_DATE tags found: {date_count}")
            
            # Return estimated record count
            estimated_records = min(eor_count, call_count, date_count) if (eor_count and call_count and date_count) else 0
            self.log(f"📊 Estimated total records: {estimated_records}")
            
            return estimated_records
            
        except Exception as e:
            self.log(f"⚠️ Error counting ADIF records: {e}")
            return 0
    
    def process_complete_qrz_data(self, adif_data):
        """
        Server-side processing van complete QRZ ADIF data
        Parseert alle records en filtert vandaag alleen
        """
        try:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.log(f"🔍 [{current_time}] === SERVER-SIDE ADIF PROCESSING ===")
            
            # Parse all ADIF records
            records = []
            lines = adif_data.split('\n')
            current_record = {}
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                    
                # Check for end of record
                if line == '<eor>':
                    if current_record:
                        records.append(current_record.copy())
                        current_record = {}
                    continue
                
                # Parse ADIF fields
                if ':' in line and '>' in line:
                    parts = line.split('>', 1)
                    if len(parts) == 2:
                        field_part = parts[0]
                        value = parts[1]
                        
                        # Extract field name and length
                        field_info = field_part.split(':')
                        if len(field_info) >= 2:
                            field_name = field_info[0]
                            # field_length = int(field_info[1])  # Not used but could be useful
                            
                            # Clean field name
                            if field_name.startswith('<'):
                                field_name = field_name[1:]
                            if field_name.endswith('>'):
                                field_name = field_name[:-1]
                                
                            current_record[field_name] = value
            
            self.log(f"📊 Parsed {len(records)} total QSO records from ADIF")
            
            # Filter voor vandaag
            today = date.today().strftime('%Y%m%d')
            today_records = []
            
            for record in records:
                qso_date = record.get('qso_date', '')
                if qso_date == today:
                    today_records.append(record)
            
            self.log(f"📅 {len(today_records)} records from today ({today})")
            
            # Sort today records by complete timestamp (newest first)
            def get_timestamp(record):
                date_str = record.get('qso_date', '00000000')
                time_str = record.get('time_on', '0000').zfill(6)  # Pad to 6 digits
                return date_str + time_str  # Format: YYYYMMDDHHMMSS
            
            today_records.sort(key=get_timestamp, reverse=True)
            self.log(f"🔄 Sorted {len(today_records)} today records by complete timestamp (newest first)")
            
            # Extract callsigns
            today_callsigns = []
            for record in today_records:
                call = record.get('call', '')
                if call:
                    # Get complete timestamp for logging
                    date_str = record.get('qso_date', '')
                    time_str = record.get('time_on', '')
                    timestamp_str = get_timestamp(record)
                    
                    today_callsigns.append({
                        'callsign': call,
                        'time': record.get('time_on', ''),
                        'mode': record.get('mode', ''),
                        'frequency': record.get('freq', ''),
                        'date': date_str,
                        'timestamp': timestamp_str,
                        'record': record
                    })
            
            self.log(f"🎯 Today callsigns found: {len(today_callsigns)}")
            
            # Log complete breakdown with timestamps
            for i, call_data in enumerate(today_callsigns, 1):
                timestamp_info = f"{call_data['date']} {call_data['time']}" if call_data['date'] and call_data['time'] else f"{call_data['timestamp']}"
                self.log(f"  {i}. {call_data['callsign']} @ {timestamp_info} ({call_data['mode']})")
            
            # Log which record will be used for PUCS comparison
            if today_callsigns:
                top_record = today_callsigns[0]
                timestamp_info = f"{top_record['date']} {top_record['time']}" if top_record['date'] and top_record['time'] else f"{top_record['timestamp']}"
                self.log(f"🎯 PUCS COMPARISON WILL USE: {top_record['callsign']} (timestamp: {timestamp_info})")
            
            # Check for missing callsigns (compared to PUCS)
            pucs_callsigns = self.get_today_active_callsigns()
            qrz_callsigns = [c['callsign'] for c in today_callsigns]
            
            missing_in_qrz = set(pucs_callsigns) - set(qrz_callsigns)
            extra_in_qrz = set(qrz_callsigns) - set(pucs_callsigns)
            
            if missing_in_qrz:
                self.log(f"⚠️ MISSING in QRZ: {list(missing_in_qrz)}")
            
            if extra_in_qrz:
                self.log(f"ℹ️ EXTRA in QRZ: {list(extra_in_qrz)}")
            
            return {
                'total_records': len(records),
                'today_records': len(today_records),
                'today_callsigns': today_callsigns,
                'missing_in_qrz': list(missing_in_qrz),
                'extra_in_qrz': list(extra_in_qrz)
            }
            
        except Exception as e:
            self.log(f"❌ Error in complete ADIF processing: {e}")
            import traceback
            self.log(f"💥 TRACEBACK: {traceback.format_exc()}")
            return None
            self.log(f"❌ Fout bij ophalen QRZ logboek: {e}")
            return None
    
    def debug_save_adif_data_enhanced(self, adif_data, context="default", metadata=None):
        """Enhanced ADIF data saving with comprehensive logging"""
        try:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            from datetime import date
            today_str = date.today().strftime('%Y%m%d')
            debug_file = f"instance/qrz_debug_enhanced_{today_str}_{context}.adif"
            
            # Comprehensive logging
            adif_preview = adif_data[:300] if len(adif_data) > 300 else adif_data
            self.log(f"🐛 [{current_time}] === ENHANCED ADIF SAVE ({context}) ===")
            self.log(f"🐛 Context: {context}")
            self.log(f"🐛 Today's date: {today_str}")
            self.log(f"🐛 Target file: {debug_file}")
            self.log(f"🐛 ADIF data length: {len(adif_data)} characters")
            self.log(f"🐛 ADIF preview: {adif_preview}...")
            
            # Log metadata if provided
            if metadata:
                self.log(f"🐛 Metadata provided:")
                for key, value in metadata.items():
                    self.log(f"   📋 {key}: {value}")
            
            # Check for key patterns
            calls_count = adif_data.lower().count('<call:')
            dates_count = adif_data.lower().count('<qso_date:')
            eor_count = adif_data.lower().count('<eor>')
            
            self.log(f"🐛 ADIF Analysis:")
            self.log(f"   📞 CALL fields: {calls_count}")
            self.log(f"   📅 QSO_DATE fields: {dates_count}")
            self.log(f"   📋 EOR markers: {eor_count}")
            
            # Ensure instance directory exists
            if not os.path.exists("instance"):
                self.log(f"🐛 Creating instance directory...")
                os.makedirs("instance")
                self.log(f"🐛 Instance directory created")
            
            # Verify write permissions
            try:
                test_file = "instance/write_test.tmp"
                with open(test_file, 'w') as f:
                    f.write("test")
                os.remove(test_file)
                self.log(f"🐛 Write permissions OK")
            except Exception as e:
                self.log(f"🐛 ❌ Write permission test failed: {e}")
                return
            
            # Write the complete ADIF data with header
            with open(debug_file, 'w', encoding='utf-8') as f:
                # Write header with metadata
                f.write(f"# PUCS QRZ ADIF Debug File\n")
                f.write(f"# Generated: {current_time}\n")
                f.write(f"# Context: {context}\n")
                f.write(f"# Data length: {len(adif_data)} chars\n")
                f.write(f"# Call fields: {calls_count}\n")
                f.write(f"# Date fields: {dates_count}\n")
                f.write(f"# EOR markers: {eor_count}\n")
                
                # Write additional metadata if provided
                if metadata:
                    f.write(f"# Additional Metadata:\n")
                    for key, value in metadata.items():
                        f.write(f"# {key}: {value}\n")
                
                f.write(f"# ======================\n\n")
                
                # Write the actual ADIF data
                written_chars = f.write(adif_data)
                f.flush()
                os.fsync(f.fileno())  # Force to disk
            
            self.log(f"🐛 Enhanced file write completed: {written_chars} chars written")
            
            # Verify file exists and has content
            if os.path.exists(debug_file):
                file_size = os.path.getsize(debug_file)
                self.log(f"🐛 ✅ SUCCESS: Enhanced debug file created!")
                self.log(f"   📂 File: {debug_file}")
                self.log(f"   📏 Size: {file_size} bytes")
                self.log(f"   📍 Path: {os.path.abspath(debug_file)}")
            else:
                self.log(f"🐛 ❌ File creation failed - file doesn't exist")
                
        except Exception as e:
            self.log(f"🐛 ❌ FATAL ERROR in debug_save_adif_data_enhanced: {type(e).__name__}: {e}")
            import traceback
            self.log(f"🐛 💥 FULL TRACEBACK:\n{traceback.format_exc()}")
            self.log(f"🐛 🔍 Current working directory: {os.getcwd()}")
            self.log(f"🐛 🔍 Directory exists check: {os.path.exists('instance')}")
            if os.path.exists('instance'):
                try:
                    files = os.listdir('instance')
                    self.log(f"🐛 📁 Instance contents: {files}")
                except Exception as e2:
                    self.log(f"🐛 ❌ Cannot list instance dir: {e2}")
    
    def debug_save_adif_data(self, adif_data):
        """Save ADIF data to file for debugging"""
        try:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # CRITICAL FIX: Ensure absolute path and force creation
            from datetime import date
            today_str = date.today().strftime('%Y%m%d')
            debug_file = f"instance/qrz_debug_{today_str}.adif"
            
            # DEBUG: Full logging before writing
            adif_preview = adif_data[:200] if len(adif_data) > 200 else adif_data
            self.log(f"🐛 [{current_time}] === STARTING ADIF SAVE ===")
            self.log(f"🐛 Today's date: {today_str}")
            self.log(f"🐛 Target file: {debug_file}")
            self.log(f"🐛 ADIF data length: {len(adif_data)} characters")
            self.log(f"🐛 ADIF preview: {adif_preview}...")
            
            # Ensure instance directory exists
            if not os.path.exists("instance"):
                self.log(f"🐛 Creating instance directory...")
                os.makedirs("instance")
                self.log(f"🐛 Instance directory created")
            else:
                self.log(f"🐛 Instance directory already exists")
            
            # Verify write permissions
            try:
                test_file = "instance/write_test.tmp"
                with open(test_file, 'w') as f:
                    f.write("test")
                os.remove(test_file)
                self.log(f"🐛 Write permissions OK")
            except Exception as e:
                self.log(f"🐛 ❌ Write permission test failed: {e}")
                return
            
            # Write the ADIF data
            with open(debug_file, 'w', encoding='utf-8') as f:
                written_chars = f.write(adif_data)
                f.flush()
                os.fsync(f.fileno())  # Force to disk
            
            self.log(f"🐛 File write completed: {written_chars} chars written")
            
            # Verify file exists and has content
            if os.path.exists(debug_file):
                file_size = os.path.getsize(debug_file)
                self.log(f"🐛 ✅ SUCCESS: Debug file created! Size: {file_size} bytes")
                self.log(f"🐛 📂 Full path: {os.path.abspath(debug_file)}")
            else:
                self.log(f"🐛 ❌ CRITICAL: File doesn't exist after write!")
                # Try to list directory contents
                try:
                    files = os.listdir("instance")
                    self.log(f"🐛 📁 Instance contents: {files}")
                except Exception as e2:
                    self.log(f"🐛 ❌ Cannot list instance dir: {e2}")
                    
        except Exception as e:
            self.log(f"🐛 ❌ FATAL ERROR in debug_save_adif_data: {type(e).__name__}: {e}")
            import traceback
            self.log(f"🐛 💥 FULL TRACEBACK:\n{traceback.format_exc()}")
            self.log(f"🐛 🔍 Current working directory: {os.getcwd()}")
            self.log(f"🐛 🔍 Directory exists check: {os.path.exists('instance')}")
            if os.path.exists('instance'):
                try:
                    self.log(f"🐛 🔍 Instance permissions: {oct(os.stat('instance').st_mode)[-3:]}")
                except Exception as perm_e:
                    self.log(f"🐛 🔍 Cannot check permissions: {perm_e}")
            
    def parse_callsigns_from_adif_today(self, adif_data):
        """Parseer callsigns uit ADIF data die vandaag een QSO hebben"""
        try:
            today = date.today()
            today_str = today.strftime('%Y%m%d')  # YYYYMMDD formaat voor ADIF
            
            self.log(f"🔍 Zoek naar QSO datum: {today_str} in ADIF data...")
            self.log(f"📊 ADIF data lengte: {len(adif_data)} karakters")
            
            # Split ADIF data in individuele QSO records (case-insensitive)
            qso_records = adif_data.replace('<eor>', '<EOR>').split('<EOR>')
            # Filter lege records
            qso_records = [record.strip() for record in qso_records if record.strip()]
            self.log(f"📊 {len(qso_records)} echte QSO records gevonden")
            
            today_callsigns = []
            all_qso_dates_found = set()
            
            for i, record in enumerate(qso_records):
                # Debug: toon eerste paar records
                if i < 3:
                    self.log(f"🔍 Record {i+1} sample: {record[:200]}...")
                
                # Zoek alleen naar QSO datum velden in ADIF
                qso_date_patterns = [
                    r'<QSO_DATE:\d+>(\d{8})',      # Standard QSO_DATE: YYYYMMDD
                    r'<QSO_DATE:\d+>(\d{4}-\d{2}-\d{2})',  # ISO format YYYY-MM-DD
                ]
                
                call_match = re.search(r'<CALL:\d+>([A-Z0-9/]+)', record, re.IGNORECASE)
                qso_date_found = None
                
                # Zoek alleen QSO datum (geen upload/download datums)
                for pattern in qso_date_patterns:
                    date_match = re.search(pattern, record, re.IGNORECASE)
                    if date_match:
                        qso_date_found = date_match.group(1)
                        break
                
                if call_match and qso_date_found:
                    callsign = call_match.group(1).upper()
                    
                    # Normaliseer QSO datum naar YYYYMMDD formaat
                    if '-' in qso_date_found:  # YYYY-MM-DD format
                        qso_date = qso_date_found.replace('-', '')
                    else:  # Al in YYYYMMDD format
                        qso_date = qso_date_found
                    
                    all_qso_dates_found.add(qso_date)
                    
                    # Debug voor eerste paar matches
                    if i < 5:
                        self.log(f"🔍 Record {i+1}: {callsign}")
                        self.log(f"   📅 QSO datum: {qso_date}")
                    
                    # Check alleen of QSO vandaag is gemaakt
                    if qso_date == today_str:
                        today_callsigns.append(callsign)
                        self.log(f"✅ Match gevonden: {callsign} (QSO vandaag: {qso_date})")
                elif call_match:
                    # Callsign gevonden maar geen QSO datum
                    callsign = call_match.group(1).upper()
                    if i < 5:
                        self.log(f"⚠️ Record {i+1}: {callsign} - geen QSO datum gevonden")
            
            # Toon alle gevonden QSO datums voor debugging
            if all_qso_dates_found:
                sorted_dates = sorted(list(all_qso_dates_found))
                self.log(f"📅 Alle QSO datums gevonden: {', '.join(sorted_dates[:10])}..." if len(sorted_dates) > 10 else f"📅 Alle QSO datums gevonden: {', '.join(sorted_dates)}")
            
            # Maak unieke lijst
            unique_callsigns = list(set(today_callsigns))
            
            if unique_callsigns:
                self.log(f"📋 {len(unique_callsigns)} unieke callsigns met QSO van vandaag: {', '.join(unique_callsigns)}")
            else:
                self.log(f"📭 Geen callsigns met QSO van vandaag ({today_str}) gevonden in logboek")
                
            return unique_callsigns
            
        except Exception as e:
            self.log(f"❌ Fout bij parsen ADIF data voor vandaag: {e}")
            return []
            
    def get_latest_callsign_from_adif(self, adif_data):
        """Haal de laatste (meest recente) callsign uit ADIF data"""
        try:
            # STAP 1: Split ADIF in individuele records
            records = re.split(r'<EOR>|<eor>', adif_data, flags=re.IGNORECASE)
            
            # STAP 2: Parse alle records met timestamps
            parsed_records = []
            
            for record in records:
                if not record.strip():
                    continue
                
                # Zoek CALL field
                call_match = re.search(r'<CALL:(\d+)>([A-Z0-9/]+)', record, re.IGNORECASE)
                if not call_match:
                    continue
                
                callsign = call_match.group(2).upper()
                
                # Zoek timestamps (optioneel)
                date_match = re.search(r'<QSO_DATE:(\d+)>(\d{8})', record, re.IGNORECASE)
                time_match = re.search(r'<TIME_ON:(\d+)>(\d{4,6})', record, re.IGNORECASE)
                
                if date_match and time_match:
                    date_str = date_match.group(2)
                    time_str = time_match.group(2).zfill(6)  # Pad to 6 digits
                    timestamp = date_str + time_str  # Format: YYYYMMDDHHMMSS
                    
                    parsed_records.append({
                        'callsign': callsign,
                        'timestamp': timestamp,
                        'date_str': date_str,
                        'time_str': time_str,
                        'record': record
                    })
                else:
                    # Geen timestamp - gebruik nul timestamp
                    parsed_records.append({
                        'callsign': callsign,
                        'timestamp': '00000000000000',
                        'date_str': 'Onbekend',
                        'time_str': 'Onbekend',
                        'record': record
                    })
            
            if not parsed_records:
                self.log("❌ Geen CALL velden gevonden in ADIF")
                return None
            
            # STAP 3: Sorteer op timestamp (laatste eerst)
            parsed_records.sort(key=lambda x: x['timestamp'], reverse=True)
            
            # STAP 4: Retourneer de laatste
            latest = parsed_records[0]
            
            if latest['timestamp'] != '00000000000000':
                self.log(f"🎯 Laatste callsign gevonden: {latest['callsign']} (datum: {latest['date_str']} {latest['time_str']})")
            else:
                self.log(f"🎯 Laatste callsign gevonden: {latest['callsign']} (geen timestamp)")
            
            # Debug: toon alle gevonden callsigns met timestamps
            self.log(f"📊 Alle callsigns gevonden ({len(parsed_records)}):")
            for i, rec in enumerate(parsed_records):
                timestamp_info = f"{rec['date_str']} {rec['time_str']}" if rec['timestamp'] != '00000000000000' else "geen timestamp"
                self.log(f"   {i+1}. {rec['callsign']} - {timestamp_info}")
            
            return latest['callsign']
                
        except Exception as e:
            self.log(f"❌ Fout bij ophalen laatste callsign uit ADIF: {e}")
            return None
    
    def get_latest_callsign_from_adif_today(self, adif_data):
        """Haal alleen de laatste (meest recente) callsign van vandaag uit ADIF data"""
        try:
            # GEBRUIK COMPLETE SERVER-SIDE PROCESSING
            self.log(f"🔍 Using complete server-side ADIF processing...")
            result = self.process_complete_qrz_data(adif_data)
            
            if not result:
                self.log("❌ Complete ADIF processing failed")
                return None
            
            today_callsigns = result['today_callsigns']
            
            if not today_callsigns:
                self.log("📭 No callsigns found for today in complete ADIF data")
                return None
            
            # Return de nieuwste callsign
            latest = today_callsigns[0]  # Already sorted newest first
            latest_callsign = latest['callsign']
            
            self.log(f"🎯 Latest callsign from complete data: {latest_callsign} @ {latest['time']}")
            self.log(f"📊 Total today callsigns: {len(today_callsigns)}")
            
            # Return in het juiste formaat voor compatibiliteit
            return latest_callsign
                
        except Exception as e:
            self.log(f"❌ Fout bij ophalen laatste callsign van vandaag uit ADIF: {e}")
            return None
            
    def check_and_remove_logged_callsigns(self):
        """Hoofdfunctie: check logboek en verwijder ALLEEN de laatst gelogde callsign van vandaag"""
        try:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.log(f"🔍 [{current_time}] === START QRZ LOGBOEK CHECK ===")
            
            # Haal QRZ configuratie op
            qrz_callsign, api_key = self.get_qrz_config()
            
            if not qrz_callsign or not api_key:
                self.log("⚠️ Geen geldige QRZ configuratie beschikbaar")
                return
            
            self.log(f"📋 QRZ Config: callsign={qrz_callsign}")
            
            # Haal actieve callsigns van vandaag op
            active_callsigns_today = self.get_today_active_callsigns()
            
            if not active_callsigns_today:
                self.log("📭 Geen actieve callsigns van vandaag om te controleren")
                self.log(f"🔍 [{current_time}] === END QRZ LOGBOEK CHECK (no active callsigns) ===")
                return
                
            self.log(f"🔍 {len(active_callsigns_today)} actieve callsigns van vandaag gevonden")
            self.log(f"🐛 Active callsigns: {active_callsigns_today}")
            
            # Haal QRZ logboek op
            adif_data = self.fetch_qrz_logbook(qrz_callsign, api_key)
            
            if not adif_data:
                self.log("❌ Kan QRZ logboek niet ophalen")
                self.log(f"🔍 [{current_time}] === END QRZ LOGBOEK CHECK (fetch failed) ===")
                return
                
            # DEBUG: Analyseer alle callsigns in ADIF data
            all_adif_callsigns = self.debug_analyze_adif_callsigns(adif_data)
                
            # Haal ALLEEN de laatste callsign van vandaag uit logboek op
            latest_logged_callsign = self.get_latest_callsign_from_adif_today(adif_data)
            
            # ALWAYS save complete ADIF data for debugging - ENHANCED VERSION
            self.debug_save_adif_data_enhanced(adif_data, f"fetch_qrz_logbook_output")
            
            if not latest_logged_callsign:
                self.log("📭 Geen gelogde callsigns van vandaag gevonden in logboek")
                self.log(f"🔍 [{current_time}] === END QRZ LOGBOEK CHECK (no logged callsigns) ===")
                return
            
            # Check of de laatst gelogde callsign voorkomt in de actieve entries van vandaag
            self.log(f"🎯 Laatst gelogde callsign van vandaag: {latest_logged_callsign}")
            
            # Kijk of er meerdere entries zijn van deze callsign van vandaag (met strings!)
            today_entries_same_callsign = [c for c in active_callsigns_today if c.upper() == latest_logged_callsign.upper()]
            
            if not today_entries_same_callsign:
                self.log(f"ℹ️ Callsign {latest_logged_callsign} niet gevonden in actieve entries van vandaag")
                self.log(f"🔍 [{current_time}] === END QRZ LOGBOEK CHECK (callign not in active) ===")
                return
            
            self.log(f"📋 {len(today_entries_same_callsign)} entry(s) van {latest_logged_callsign} van vandaag gevonden")
            
            # Verwijder ALLEEN de laatste entry van deze callsign (niet alle)
            if self.remove_callsign_from_db(latest_logged_callsign):
                self.log(f"🎯 Laatste entry van {latest_logged_callsign} verwijderd uit entry list")
                self.log(f"🔍 [{current_time}] === END QRZ LOGBOEK CHECK (success) ===")
            else:
                self.log(f"⚠️ Kon {latest_logged_callsign} niet verwijderen uit database")
                self.log(f"🔍 [{current_time}] === END QRZ LOGBOEK CHECK (removal failed) ===")
                
        except Exception as e:
            self.log(f"❌ Fout tijdens logboek check: {e}")
            self.log(f"🔍 === ERROR END QRZ LOGBOEK CHECK ===")
            
    def check_loop(self):
        """Hoofdloop die elke minuut draait"""
        self.log(f"🚀 QRZ Logbook Checker gestart (interval: {self.interval//60} min, alleen vandaag)")
        
        while self.running:
            try:
                self.check_and_remove_logged_callsigns()
                
                # Wacht interval seconden, maar check elke 5 seconden of we moeten stoppen
                for i in range(0, self.interval, 5):
                    if not self.running:
                        break
                    time.sleep(5)
                    
            except KeyboardInterrupt:
                self.log("🛑 Onderbroken door gebruiker")
                break
            except Exception as e:
                self.log(f"❌ Onverwachte fout in hoofdloop: {e}")
                time.sleep(5)  # Korte pauze bij fouten
                
        self.log("🛑 QRZ Logbook Checker gestopt")
        
    def start(self):
        """Start de checker service"""
        if self.running:
            self.log("⚠️ Service draait al")
            return
            
        self.running = True
        self.check_thread = threading.Thread(target=self.check_loop, daemon=True)
        self.check_thread.start()
        self.log("✅ Service gestart als background thread")
        
    def stop(self):
        """Stop de checker service"""
        if not self.running:
            self.log("⚠️ Service draait niet")
            return
            
        self.running = False
        if self.check_thread:
            self.check_thread.join(timeout=5)
        self.log("✅ Service gestopt")
        
    def is_running(self):
        """Check of service draait"""
        return self.running and self.check_thread and self.check_thread.is_alive()
        
    def decode_html_entities(self, text):
        """Decode HTML entities in text"""
        import html
        return html.unescape(text)
    
    def debug_analyze_adif_callsigns(self, adif_data):
        """Analyseer alle callsigns in ADIF data voor debugging"""
        try:
            records = re.split(r'<EOR>|<eor>', adif_data, flags=re.IGNORECASE)
            all_callsigns = []
            
            for i, record in enumerate(records):
                if not record.strip():
                    continue
                
                # Zoek CALL field
                call_match = re.search(r'<CALL:(\d+)>([A-Z0-9/]+)', record, re.IGNORECASE)
                if call_match:
                    callsign = call_match.group(2).upper()
                    
                    # Zoek QSO datum
                    date_match = re.search(r'<QSO_DATE:(\d+)>(\d{8})', record, re.IGNORECASE)
                    if date_match:
                        qso_date = date_match.group(2)
                        all_callsigns.append({
                            'callsign': callsign,
                            'qso_date': qso_date,
                            'record_index': i + 1
                        })
            
            self.log(f"🔍 ADIF Analysis - Total QSO records: {len(all_callsigns)}")
            
            # Groepeer per datum
            by_date = {}
            for cs in all_callsigns:
                date = cs['qso_date']
                if date not in by_date:
                    by_date[date] = []
                by_date[date].append(cs)
            
            for date, callsigns in sorted(by_date.items()):
                self.log(f"📅 {date}: {len(callsigns)} callsigns - {[cs['callsign'] for cs in callsigns[:3]]}" + 
                        ("..." if len(callsigns) > 3 else ""))
            
            return all_callsigns
            
        except Exception as e:
            self.log(f"❌ Fout bij ADIF analysis: {e}")
            return []

# Global QRZ checker instance
qrz_checker = QRZLogbookChecker()

class QRZLatestCallsignMonitor:
    """
    Monitor voor de laatste ingevoerde callsign in QRZ logboek.
    Checkt elke minuut de laatste callsign en verwijdert deze automatisch uit PUCS als gevonden.
    """
    def __init__(self, db_path='instance/radio_entry.db'):
        self.db_path = db_path
        self.running = False
        self.monitor_thread = None
        self.interval = 60  #  minuten in seconden
        self.last_checked_callsign = None  # Om herhaalde checks te voorkomen
        
    def log(self, message):
        """Log met timestamp"""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"[{timestamp}] QRZ Latest Monitor: {message}")
        
    @contextmanager
    def get_db_connection(self):
        """Database connectie context manager"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
        finally:
            conn.close()
            
    def get_qrz_config(self):
        """Haal QRZ configuratie op uit database"""
        try:
            with self.get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT callsign, api_key FROM qrz_config ORDER BY updated_at DESC LIMIT 1")
                row = cursor.fetchone()
                
                if row:
                    return row['callsign'], row['api_key']
                else:
                    # Fallback naar default waarden
                    self.log("Geen QRZ configuratie gevonden, gebruik default waarden")
                    return 'PH25XMAS', '4DB9-.....-579F'
                    
        except Exception as e:
            self.log(f"Fout bij ophalen QRZ config: {e}")
            return 'PH25XMAS', '4DB9-.....-579F'
            
    def check_callsign_in_pucs(self, callsign):
        """Check of callsign bestaat in PUCS entries"""
        try:
            with self.get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM callsign_entry WHERE callsign = ?", (callsign,))
                row = cursor.fetchone()
                return row is not None
                
        except Exception as e:
            self.log(f"Fout bij checken callsign {callsign} in PUCS: {e}")
            return False
            
    def remove_callsign_from_pucs(self, callsign):
        """Verwijder callsign uit PUCS entries"""
        try:
            with self.get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM callsign_entry WHERE callsign = ?", (callsign,))
                if cursor.rowcount > 0:
                    conn.commit()
                    self.log(f"🗑️ Callsign {callsign} automatisch verwijderd uit PUCS (was gelogd in QRZ)")
                    return True
                return False
                
        except Exception as e:
            self.log(f"Fout bij verwijderen callsign {callsign}: {e}")
            return False
            
    def monitor_latest_callsign(self):
        """
        Monitor functie: haal laatste callsign uit QRZ en check tegen PUCS
        """
        try:
            # Haal QRZ configuratie op
            qrz_callsign, api_key = self.get_qrz_config()
            
            if not qrz_callsign or not api_key:
                self.log("⚠️ Geen geldige QRZ configuratie beschikbaar")
                return False
                
            self.log(f"🔍 Monitor QRZ logboek van {qrz_callsign} voor laatste callsign...")
            
            # Haal QRZ logboek op via de bestaande checker
            adif_data = qrz_checker.fetch_qrz_logbook(qrz_callsign, api_key)
            
            if not adif_data:
                self.log("❌ Kan QRZ logboek niet ophalen")
                return False
                
            # Haal laatste callsign op
            latest_callsign = qrz_checker.get_latest_callsign_from_adif(adif_data)
            
            if not latest_callsign:
                self.log("📭 Geen laatste callsign gevonden in QRZ logboek")
                return False
                
            # Skip als we deze callsign al gecheckt hebben
            if latest_callsign == self.last_checked_callsign:
                return True
                
            self.log(f"🎯 Laatste QRZ callsign: {latest_callsign}")
            
            # Check of callsign bestaat in PUCS
            if self.check_callsign_in_pucs(latest_callsign):
                self.log(f"✅ Callsign {latest_callsign} gevonden in PUCS, verwijderen...")
                if self.remove_callsign_from_pucs(latest_callsign):
                    self.last_checked_callsign = latest_callsign
                    return True
                else:
                    self.log(f"❌ Kon callsign {latest_callsign} niet verwijderen uit PUCS")
            else:
                self.log(f"ℹ️ Callsign {latest_callsign} niet gevonden in PUCS, wachten...")
                
            # Update last checked callsign
            self.last_checked_callsign = latest_callsign
            return True
            
        except Exception as e:
            self.log(f"❌ Fout tijdens monitoren laatste callsign: {e}")
            return False
            
    def monitor_loop(self):
        """Hoofdmonitor loop - draait continu"""
        self.log(f"🚀 QRZ Latest Callsign Monitor gestart (interval: {self.interval//60} min)")
        
        while self.running:
            try:
                # Voer monitor check uit
                self.monitor_latest_callsign()
                
                # Wacht interval seconden, maar check elke 5 seconden of we moeten stoppen
                for i in range(0, self.interval, 5):
                    if not self.running:
                        break
                    time.sleep(5)
                    
            except KeyboardInterrupt:
                self.log("🛑 Onderbroken door gebruiker")
                break
            except Exception as e:
                self.log(f"❌ Onverwachte fout in monitor loop: {e}")
                time.sleep(5)  # Korte pauze bij fouten
                
        self.log("🛑 QRZ Latest Callsign Monitor gestopt")
        
    def start(self):
        """Start de monitor service"""
        if self.running:
            self.log("⚠️ Monitor service draait al")
            return
            
        self.running = True
        self.monitor_thread = threading.Thread(target=self.monitor_loop, daemon=True)
        self.monitor_thread.start()
        self.log("✅ Monitor service gestart als background thread")
        
    def stop(self):
        """Stop de monitor service"""
        if not self.running:
            self.log("⚠️ Monitor service draait niet")
            return
            
        self.running = False
        if self.monitor_thread:
            self.monitor_thread.join(timeout=5)
        self.log("✅ Monitor service gestopt")
        
    def is_running(self):
        """Check of service draait"""
        return self.running and self.monitor_thread and self.monitor_thread.is_alive()

# Global QRZ latest monitor instance
qrz_latest_monitor = QRZLatestCallsignMonitor()

def is_local_request():
    client_ip = request.environ.get('HTTP_X_REAL_IP', request.remote_addr)
    return (
        client_ip in ['127.0.0.1', 'localhost']
        or client_ip.startswith('192.168.')
        or client_ip.startswith('10.')
        or client_ip.startswith('172.')
    )

def create_empty_excel_file():
    """
    Maak een leeg Excel-bestand aan met alleen headers.
    """
    try:
        excel_file_path = os.path.join('instance', 'radio_entries_log.xlsx')
        
        # Maak nieuwe workbook met headers
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Radio Entries Log"
        
        # Voeg headers toe
        headers = ['Callsign', 'QTH', 'Remarks', 'Timestamp']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Stel kolombreedte in
        worksheet.column_dimensions['A'].width = 15  # Callsign
        worksheet.column_dimensions['B'].width = 25  # QTH
        worksheet.column_dimensions['C'].width = 35  # Remarks
        worksheet.column_dimensions['D'].width = 20  # Timestamp
        
        # Sla de workbook op
        workbook.save(excel_file_path)
        print(f"📊 Leeg Excel bestand aangemaakt: {excel_file_path}")
        return True
        
    except Exception as e:
        print(f"❌ Fout bij aanmaken leeg Excel bestand: {e}")
        return False

def log_entry_to_excel(callsign, qth, remarks, timestamp=None):
    """
    Log een nieuwe entry naar Excel-file in de instance map.
    
    Args:
        callsign (str): De callsign van de entry
        qth (str): Locatie/QTH van de entry  
        remarks (str): Opmerkingen/remarks van de entry
        timestamp (datetime): Timestamp van de entry (standaard: nu)
    """
    try:
        if timestamp is None:
            timestamp = datetime.now()
        
        # Pad naar de Excel-file in de instance map
        excel_file_path = os.path.join('instance', 'radio_entries_log.xlsx')
        
        # Check of het bestand al bestaat
        if os.path.exists(excel_file_path):
            # Laad bestaande workbook
            workbook = load_workbook(excel_file_path)
            worksheet = workbook.active
        else:
            # Maak nieuwe workbook met headers
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Radio Entries Log"
            
            # Voeg headers toe
            headers = ['Callsign', 'QTH', 'Remarks', 'Timestamp']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Stel kolombreedte in
            worksheet.column_dimensions['A'].width = 15  # Callsign
            worksheet.column_dimensions['B'].width = 25  # QTH
            worksheet.column_dimensions['C'].width = 35  # Remarks
            worksheet.column_dimensions['D'].width = 20  # Timestamp
        
        # Vind de volgende lege rij
        next_row = worksheet.max_row + 1
        
        # Voeg de data toe
        worksheet.cell(row=next_row, column=1, value=callsign or "")
        worksheet.cell(row=next_row, column=2, value=qth or "")
        worksheet.cell(row=next_row, column=3, value=remarks or "")
        worksheet.cell(row=next_row, column=4, value=timestamp.strftime("%Y-%m-%d %H:%M:%S"))
        
        # Sla de workbook op
        workbook.save(excel_file_path)
        
        print(f"📊 Entry gelogd naar Excel: {callsign} op {timestamp.strftime('%Y-%m-%d %H:%M:%S')}")
        return True
        
    except Exception as e:
        print(f"❌ Fout bij Excel logging: {e}")
        return False

# =============================================================================
# PUBLIC API ROUTES (CORS enabled voor frontend)
# =============================================================================

@app.route('/api/entries', methods=['GET', 'OPTIONS'])
def api_get_entries():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        config = Config.get_current()
        entries = CallsignEntry.query.order_by(CallsignEntry.position).all()
        entry_positions = {i: None for i in range(1, 5)}
        for entry in entries:
            entry_positions[entry.position] = entry.callsign

        submit_position = None
        for pos in range(1, 5):
            if entry_positions[pos] is None:
                submit_position = pos
                break

        return jsonify({
            'config': {
                'operator_name': config.operator_name,
                'frequency': config.frequency
            },
            'entries': entry_positions,
            'submit_position': submit_position,
            'timestamp': datetime.now().isoformat()
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/submit_callsign', methods=['POST', 'OPTIONS'])
def api_submit_callsign():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Geen JSON data ontvangen'}), 400

        callsign = data.get('callsign', '').strip().upper()
        location = data.get('location', '').strip()  # QTH toegevoegd
        comment = data.get('comment', '').strip()    # Remarks toegevoegd

        if not callsign:
            return jsonify({'error': 'Callsign is verplicht'}), 400

        if not callsign.replace('/', '').replace('-', '').isalnum():
            return jsonify({'error': 'Ongeldige callsign format'}), 400

        if len(callsign) < 3:
            return jsonify({'error': 'Callsign te kort (minimum 3 karakters)'}), 400

        occupied_positions = [entry.position for entry in CallsignEntry.query.all()]
        available_position = None
        for pos in range(1, 7):  # van 1,5 naar 1,7 veranderd
            if pos not in occupied_positions:
                available_position = pos
                break

        if available_position is None:
            return jsonify({'error': 'Alle posities zijn bezet'}), 400

        existing = CallsignEntry.query.filter_by(callsign=callsign).first()
        if existing:
            return jsonify({'error': 'Deze callsign staat al in de lijst'}), 400

        # Voeg nieuwe entry toe met alle velden
        entry = CallsignEntry(
            position=available_position, 
            callsign=callsign,
            location=location if location else None,
            comment=comment if comment else None
        )
        db.session.add(entry)
        db.session.commit()

        # Log de entry naar Excel
        log_entry_to_excel(
            callsign=callsign,
            qth=location,
            remarks=comment,
            timestamp=datetime.now()
        )

        print(f"🔄 Nieuwe entry ontvangen:")
        print(f"  - Callsign: {callsign}")
        print(f"  - QTH: {location}")
        print(f"  - Remarks: {comment}")
        print(f"✅ Entry opgeslagen in database")
        print(f"✅ Nieuwe entry: {callsign} op positie {available_position}")
        if location:
            print(f"   📍 QTH: {location}")
        if comment:
            print(f"   💬 Remarks: {comment}")

        return jsonify({
            'success': True,
            'position': available_position,
            'callsign': callsign,
            'location': location,
            'comment': comment,
            'timestamp': datetime.now().isoformat()  # Fixed deprecated datetime.utcnow()
        })

    except Exception as e:
        print(f"❌ Submit error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/get_active_callsigns', methods=['GET'])
def api_get_active_callsigns():
    try:
        entries = CallsignEntry.query.all()
        callsigns = [entry.callsign for entry in entries]
        # socketio.emit('entries_updated')  # Commented out - socketio not imported
        return jsonify({
            'callsigns': callsigns,
            'count': len(callsigns),
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clear_callsign', methods=['POST'])
def api_clear_callsign():
    try:
        data = request.get_json()
        callsign = data.get('callsign', '').strip().upper()
        api_key = data.get('api_key', '')

        expected_key = os.environ.get('API_KEY', 'logboek-script-2025')
        if api_key != expected_key:
            return jsonify({'error': 'Ongeldige API key'}), 401

        if not callsign:
            return jsonify({'error': 'Callsign is verplicht'}), 400

        entry = CallsignEntry.query.filter_by(callsign=callsign).first()
        if entry:
            position = entry.position
            db.session.delete(entry)
            db.session.commit()
            print(f"🧹 Callsign {callsign} verwijderd van positie {position} (via logboek)")
            # socketio.emit('entries_updated')  # Commented out - socketio not imported
            return jsonify({
                'success': True,
                'message': f'Callsign {callsign} verwijderd van positie {position}',
                'timestamp': datetime.now().isoformat()
            })

        return jsonify({
            'success': False,
            'message': f'Callsign {callsign} niet gevonden'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# =============================================================================
# ADMIN API ROUTES (CORS enabled voor frontend)
# =============================================================================

@app.route('/api/admin/login', methods=['POST', 'OPTIONS'])
def api_admin_login():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Geen JSON data ontvangen'}), 400

        password = data.get('password', '').strip()

        if not password:
            return jsonify({'error': 'Wachtwoord is verplicht'}), 400

        admin_password = os.environ.get('ADMIN_PASSWORD', '<password>')

        if password == admin_password:
            import hashlib
            token = hashlib.md5(f"{password}-{datetime.now().isoformat()}".encode()).hexdigest()
            print(f"✅ Admin login via API geaccepteerd")
            return jsonify({
                'success': True,
                'token': token,
                'message': 'Admin ingelogd',
                'timestamp': datetime.now().isoformat()
            })
        else:
            print(f"❌ Admin login via API mislukt: verkeerd wachtwoord")
            return jsonify({'error': 'Ongeldig wachtwoord'}), 401

    except Exception as e:
        print(f"❌ Admin login API error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/update_header', methods=['POST', 'OPTIONS'])
def api_admin_update_header():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Geen geldige autorisatie'}), 401

        token = auth_header.split(' ')[1]
        if not token:
            return jsonify({'error': 'Ongeldige token'}), 401

        data = request.get_json()
        if not data:
            return jsonify({'error': 'Geen JSON data ontvangen'}), 400

        operator_name = data.get('operator_name', '').strip()
        frequency = data.get('frequency', '').strip()

        if not operator_name or not frequency:
            return jsonify({'error': 'Operator naam en frequentie zijn verplicht'}), 400

        config = Config.get_current()
        config.operator_name = operator_name
        config.frequency = frequency
        config.updated_at = datetime.now()

        db.session.commit()

        print(f"🔧 Station info bijgewerkt via API: {operator_name} @ {frequency}")
        # socketio.emit('entries_updated')  # Commented out - socketio not imported

        return jsonify({
            'success': True,
            'message': 'Station info bijgewerkt',
            'config': {
                'operator_name': config.operator_name,
                'frequency': config.frequency
            },
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        print(f"❌ Update header API error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/delete/<int:position>', methods=['POST', 'OPTIONS'])
def api_admin_delete_entry(position):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Geen geldige autorisatie'}), 401

        token = auth_header.split(' ')[1]
        if not token:
            return jsonify({'error': 'Ongeldige token'}), 401

        if position < 1 or position > 4:
            return jsonify({'error': 'Ongeldige positie (1-4)'}), 400

        entry = CallsignEntry.query.filter_by(position=position).first()
        if entry:
            callsign = entry.callsign
            db.session.delete(entry)
            db.session.commit()
            print(f"🗑️ Entry verwijderd via API: {callsign} (positie {position})")
            # socketio.emit('entries_updated')  # Commented out - socketio not imported
            return jsonify({
                'success': True,
                'message': f'Entry {callsign} op positie {position} verwijderd',
                'position': position,
                'callsign': callsign,
                'timestamp': datetime.now().isoformat()
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Geen entry gevonden op positie {position}'
            })
    except Exception as e:
        print(f"❌ Delete entry API error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/qrz_status', methods=['GET', 'OPTIONS'])
def api_admin_qrz_status():
    """QRZ Checker status endpoint"""
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Geen geldige autorisatie'}), 401

        token = auth_header.split(' ')[1]
        if not token:
            return jsonify({'error': 'Ongeldige token'}), 401

        return jsonify({
            'success': True,
            'qrz_checker': {
                'running': qrz_checker.is_running(),
                'interval': qrz_checker.interval,
                'status': 'Actief' if qrz_checker.is_running() else 'Gestopt'
            },
            'qrz_latest_monitor': {
                'running': qrz_latest_monitor.is_running(),
                'interval': qrz_latest_monitor.interval,
                'last_checked_callsign': qrz_latest_monitor.last_checked_callsign,
                'status': 'Actief' if qrz_latest_monitor.is_running() else 'Gestopt'
            },
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        print(f"❌ QRZ status API error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/qrz_start', methods=['POST', 'OPTIONS'])
def api_admin_qrz_start():
    """Start QRZ Checker service"""
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Geen geldige autorisatie'}), 401

        token = auth_header.split(' ')[1]
        if not token:
            return jsonify({'error': 'Ongeldige token'}), 401

        if qrz_checker.is_running():
            return jsonify({
                'success': False,
                'message': 'QRZ Checker draait al',
                'status': 'Actief',
                'timestamp': datetime.now().isoformat()
            })

        qrz_checker.start()
        return jsonify({
            'success': True,
            'message': 'QRZ Checker gestart',
            'status': 'Actief',
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        print(f"❌ QRZ start API error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/qrz_stop', methods=['POST', 'OPTIONS'])
def api_admin_qrz_stop():
    """Stop QRZ Checker service"""
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Geen geldige autorisatie'}), 401

        token = auth_header.split(' ')[1]
        if not token:
            return jsonify({'error': 'Ongeldige token'}), 401

        if not qrz_checker.is_running():
            return jsonify({
                'success': False,
                'message': 'QRZ Checker draait niet',
                'status': 'Gestopt',
                'timestamp': datetime.now().isoformat()
            })

        qrz_checker.stop()
        return jsonify({
            'success': True,
            'message': 'QRZ Checker gestopt',
            'status': 'Gestopt',
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        print(f"❌ QRZ stop API error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/qrz_latest_start', methods=['POST', 'OPTIONS'])
def api_admin_qrz_latest_start():
    """Start QRZ Latest Callsign Monitor service"""
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Geen geldige autorisatie'}), 401

        token = auth_header.split(' ')[1]
        if not token:
            return jsonify({'error': 'Ongeldige token'}), 401

        if qrz_latest_monitor.is_running():
            return jsonify({
                'success': False,
                'message': 'QRZ Latest Monitor draait al',
                'status': 'Actief',
                'timestamp': datetime.now().isoformat()
            })

        qrz_latest_monitor.start()
        return jsonify({
            'success': True,
            'message': 'QRZ Latest Callsign Monitor gestart',
            'status': 'Actief',
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        print(f"❌ QRZ latest start API error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/qrz_latest_stop', methods=['POST', 'OPTIONS'])
def api_admin_qrz_latest_stop():
    """Stop QRZ Latest Callsign Monitor service"""
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Geen geldige autorisatie'}), 401

        token = auth_header.split(' ')[1]
        if not token:
            return jsonify({'error': 'Ongeldige token'}), 401

        if not qrz_latest_monitor.is_running():
            return jsonify({
                'success': False,
                'message': 'QRZ Latest Monitor draait niet',
                'status': 'Gestopt',
                'timestamp': datetime.now().isoformat()
            })

        qrz_latest_monitor.stop()
        return jsonify({
            'success': True,
            'message': 'QRZ Latest Callsign Monitor gestopt',
            'status': 'Gestopt',
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        print(f"❌ QRZ latest stop API error: {e}")
        return jsonify({'error': str(e)}), 500

# =============================================================================
# ADMIN ROUTES (alleen lokale toegang)
# =============================================================================

@app.route('/')
def index():
    return redirect(url_for('admin_login'))

@app.route('/admin')
def admin_login():
    if 'admin_id' in session:
        return redirect(url_for('admin_dashboard'))
    return render_template('admin_login.html')

@app.route('/admin/login', methods=['POST'])
def admin_login_post():
    username = request.form.get('username')
    password = request.form.get('password')

    admin = Admin.query.filter_by(username=username).first()

    if admin and admin.check_password(password):
        session['admin_id'] = admin.id
        flash('Succesvol ingelogd!', 'success')
        return redirect(url_for('admin_dashboard'))

    flash('Ongeldige gebruikersnaam of wachtwoord', 'error')
    return redirect(url_for('admin_login'))

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_id', None)
    flash('Uitgelogd', 'info')
    return redirect(url_for('admin_login'))

@app.route('/admin/download_excel')
def download_excel_entries():
    """Download het Excel-bestand met alle entries"""
    if 'admin_id' not in session:
        flash('Je moet ingelogd zijn om het Excel-bestand te downloaden', 'error')
        return redirect(url_for('admin_login'))
    
    try:
        excel_file_path = os.path.join('instance', 'radio_entries_log.xlsx')
        
        # Check of het bestand bestaat
        if not os.path.exists(excel_file_path):
            # Als het bestand niet bestaat, maak een leeg bestand aan met headers
            create_empty_excel_file()
            
        # Download het bestand
        return send_file(
            excel_file_path,
            as_attachment=True,
            download_name=f'radio_entries_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Fout bij downloaden Excel bestand: {e}")
        flash('Er is een fout opgetreden bij het downloaden van het Excel-bestand', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/dashboard')
def admin_dashboard():
    if 'admin_id' not in session:
        return redirect(url_for('admin_login'))

    config = Config.get_current()
    entries = CallsignEntry.query.order_by(CallsignEntry.position).all()
    return render_template('admin_dashboard.html', config=config, entries=entries)

@app.route('/admin/update_config', methods=['POST'])
def admin_update_config():
    if 'admin_id' not in session:
        return jsonify({'error': 'Niet geautoriseerd'}), 401

    config = Config.get_current()
    config.operator_name = request.form.get('operator_name', config.operator_name)
    config.frequency = request.form.get('frequency', config.frequency)
    config.updated_at = datetime.now()

    db.session.commit()
    flash('Configuratie bijgewerkt!', 'success')
    print(f"🔧 Configuratie bijgewerkt: {config.operator_name} @ {config.frequency}")
    # socketio.emit('entries_updated')  # Commented out - socketio not imported
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete_entry/<int:position>', methods=['POST'])
def admin_delete_entry(position):
    if 'admin_id' not in session:
        return jsonify({'error': 'Niet geautoriseerd'}), 401

    entry = CallsignEntry.query.filter_by(position=position).first()
    if entry:
        callsign = entry.callsign
        db.session.delete(entry)
        db.session.commit()
        flash(f'Entry {callsign} op positie {position} verwijderd', 'success')
        print(f"🗑️ Entry verwijderd: {callsign} (positie {position})")
        # socketio.emit('entries_updated')  # Commented out - socketio not imported
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/clear_all', methods=['POST'])
def admin_clear_all():
    if 'admin_id' not in session:
        return jsonify({'error': 'Niet geautoriseerd'}), 401

    count = CallsignEntry.query.count()
    CallsignEntry.query.delete()
    db.session.commit()
    flash(f'{count} entries gewist', 'success')
    print(f"🧹 Alle entries gewist ({count} items)")
    # socketio.emit('entries_updated')  # Commented out - socketio not imported
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/qrz_service_status', methods=['GET'])
def admin_qrz_service_status():
    """Get QRZ service status"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Niet geautoriseerd'}), 401
    
    return jsonify({
        'running': qrz_checker.running,
        'message': 'QRZ checker actief' if qrz_checker.running else 'QRZ checker gestopt'
    })

@app.route('/admin/get_qrz_config', methods=['GET'])
def admin_get_qrz_config():
    """Get current QRZ configuration"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Niet geautoriseerd'}), 401
    
    try:
        config = QRZConfig.query.order_by(QRZConfig.updated_at.desc()).first()
        if config:
            return jsonify({
                'callsign': config.callsign,
                'api_key': config.api_key,
                'success': True
            })
        else:
            # Return default values
            return jsonify({
                'callsign': 'PH25XMAS',
                'api_key': '4DB9-.....-579F',
                'success': True
            })
    except Exception as e:
        return jsonify({'error': f'Fout bij ophalen configuratie: {str(e)}', 'success': False}), 500

@app.route('/admin/qrz_config', methods=['POST'])
def admin_update_qrz_config():
    """Update QRZ configuration"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Niet geautoriseerd'}), 401
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Geen data ontvangen', 'success': False}), 400
        
        callsign = data.get('callsign', '').strip().upper()
        api_key = data.get('api_key', '').strip()
        
        if not callsign or not api_key:
            return jsonify({'error': 'Callsign en API key zijn verplicht', 'success': False}), 400
        
        # Maak nieuwe configuratie entry
        new_config = QRZConfig(
            callsign=callsign,
            api_key=api_key,
            updated_at=datetime.now()
        )
        
        db.session.add(new_config)
        db.session.commit()
        
        print(f"🔧 QRZ configuratie bijgewerkt: {callsign}")
        
        return jsonify({
            'success': True,
            'message': f'QRZ configuratie bijgewerkt voor {callsign}',
            'callsign': callsign
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Fout bij bijwerken QRZ configuratie: {e}")
        return jsonify({'error': f'Fout bij opslaan: {str(e)}', 'success': False}), 500

@app.route('/admin/test_qrz_api', methods=['POST'])
def admin_test_qrz_api():
    """Test QRZ API verbinding"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Niet geautoriseerd'}), 401
    
    try:
        # Haal huidige QRZ configuratie op
        config = QRZConfig.query.order_by(QRZConfig.updated_at.desc()).first()
        if not config:
            return jsonify({
                'success': False, 
                'error': 'Geen QRZ configuratie gevonden',
                'test_result': '❌ Geen configuratie'
            }), 404
        
        callsign = config.callsign
        api_key = config.api_key
        
        # Test QRZ API verbinding
        test_result = qrz_checker.fetch_qrz_logbook(callsign, api_key)
        
        if test_result:
            return jsonify({
                'success': True,
                'callsign': callsign,
                'api_key': api_key[:8] + '...' + api_key[-4:] if len(api_key) > 12 else '***',
                'test_result': '✅ API verbinding succesvol',
                'response_length': len(test_result),
                'message': f'QRZ logboek succesvol opgehaald voor {callsign}'
            })
        else:
            return jsonify({
                'success': False,
                'callsign': callsign,
                'api_key': api_key[:8] + '...' + api_key[-4:] if len(api_key) > 12 else '***',
                'test_result': '❌ API verbinding mislukt',
                'error': 'Kan QRZ logboek niet ophalen - controleer callsign en API key'
            })
            
    except Exception as e:
        print(f"❌ Fout bij testen QRZ API: {e}")
        return jsonify({
            'success': False,
            'test_result': '❌ Test gefaald',
            'error': f'Fout bij testen: {str(e)}'
        }), 500

# =============================================================================
# STATUS EN INFO ROUTES
# =============================================================================

@app.route('/status')
def status():
    return jsonify({
        'status': 'running',
        'version': '1.0.0',
        'timestamp': datetime.now().isoformat(),
        'entries_count': CallsignEntry.query.count(),
        'is_local': is_local_request(),
        'qrz_checker': {
            'running': qrz_checker.is_running(),
            'interval': qrz_checker.interval,
            'status': 'Actief' if qrz_checker.is_running() else 'Gestopt'
        }
    })

@app.route('/config')
def get_config():
    config = Config.get_current()
    return jsonify({
        'operator_name': config.operator_name,
        'frequency': config.frequency,
        'updated_at': config.updated_at.isoformat()
    })

# =============================================================================
# ERROR HANDLERS
# =============================================================================

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Endpoint niet gevonden'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': 'Server fout'}), 500

# =============================================================================
# STARTUP FUNCTIONS
# =============================================================================

def create_tables():
    with app.app_context():
        db.create_all()
        if not Admin.query.filter_by(username='pa3efr').first():
            admin = Admin(username='pa3efr')
            admin.set_password('<password>')  # ← HIER kun je het hardcoded wachtwoord wijzigen
            db.session.add(admin)
            db.session.commit()
            print("✅ Default admin account aangemaakt.")
        if not Config.query.first():
            config = Config(operator_name='PA3EFR', frequency='145.500 MHz')
            db.session.add(config)
            db.session.commit()
            print("✅ Default configuratie aangemaakt")
        if not QRZConfig.query.first():
            qrz_config = QRZConfig(
                callsign='PH25XMAS',
                api_key='4DB9-.....-579F'
            )
            db.session.add(qrz_config)
            db.session.commit()
            print("✅ Default QRZ configuratie aangemaakt")

def start_qrz_checker():
    """Start de QRZ checker service"""
    try:
        qrz_checker.start()
        print("✅ QRZ Logbook Checker gestart")
    except Exception as e:
        print(f"❌ Fout bij starten QRZ checker: {e}")

def stop_qrz_checker():
    """Stop de QRZ checker service"""
    try:
        qrz_checker.stop()
        print("✅ QRZ Logbook Checker gestopt")
    except Exception as e:
        print(f"❌ Fout bij stoppen QRZ checker: {e}")

def start_latest_callsign_monitor():
    """Start QRZ latest callsign monitor service"""
    try:
        qrz_latest_monitor.start()
        print("✅ QRZ Latest Callsign Monitor gestart")
    except Exception as e:
        print(f"❌ Fout bij starten QRZ latest monitor: {e}")

def stop_latest_callsign_monitor():
    """Stop QRZ latest callsign monitor service"""
    try:
        qrz_latest_monitor.stop()
        print("✅ QRZ Latest Callsign Monitor gestopt")
    except Exception as e:
        print(f"❌ Fout bij stoppen QRZ latest monitor: {e}")

def signal_handler(sig, frame):
    """Signal handler voor graceful shutdown"""
    print('\n🛑 Shutdown signaal ontvangen...')
    stop_qrz_checker()
    stop_latest_callsign_monitor()
    print('🛑 Server wordt afgesloten...')
    sys.exit(0)

def get_local_ip():
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

if __name__ == '__main__':
    # Registreer signal handlers voor graceful shutdown
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)
    
    # Registreer atexit handlers als backup
    atexit.register(stop_qrz_checker)
    atexit.register(stop_latest_callsign_monitor)
    
    # Initialiseer database
    create_tables()
    
    # Start beide QRZ services
    start_qrz_checker()
    start_latest_callsign_monitor()
    
    local_ip = get_local_ip()
    port = int(os.environ.get('PORT', 5000))
    print("\n" + "="*60)
    print("🎙️ Pile-Up Control System - Backend Server")
    print("="*60)
    print(f"🖥️ Server adres: http://{local_ip}:{port}")
    print(f"🔐 Admin interface: http://127.0.0.1:{port}/admin")
    print(f"📡 API voor frontend: http://{local_ip}:{port}/api/entries")
    print(f"⚙️ Status endpoint: http://{local_ip}:{port}/status")
    print("="*60)
    print("📝 FRONTEND CONFIGURATIE:")
    print(f"   Wijzig in frontend.js: apiBaseUrl: 'http://{local_ip}:{port}'")
    print("   Upload frontend bestanden naar pucs.pa3efr.nl")
    print("="*60)
    print("🔍 QRZ CHECKER:")
    print(f"   QRZ Checker status: {'Actief' if qrz_checker.is_running() else 'Gestopt'}")
    print(f"   Check interval: {qrz_checker.interval} seconden")
    print(f"   Controleert alle entries van vandaag")
    print("   QRZ Latest Monitor: {'Actief' if qrz_latest_monitor.is_running() else 'Gestopt'}")
    print(f"   Monitor interval: {qrz_latest_monitor.interval} seconden")
    print(f"   Monitort laatste QRZ callsign en verwijdert automatisch uit PUCS")
    print("="*60)
    print("🚀 Server gestart... Druk CTRL+C om te stoppen")
    print()
    
    try:
        socketio.run(app, host='0.0.0.0', port=port)
    except KeyboardInterrupt:
        print('\n🛑 Keyboard interrupt ontvangen')
        stop_qrz_checker()
        stop_latest_callsign_monitor()
    except Exception as e:
        print(f'\n❌ Server fout: {e}')
        stop_qrz_checker()
        stop_latest_callsign_monitor()
    finally:
        print('🛑 Server gestopt')
 #   app.run(debug=False, host='0.0.0.0', port=port)